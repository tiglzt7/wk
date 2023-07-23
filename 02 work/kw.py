import os
import numpy as np
import pandas as pd
import tkinter as tk
import openpyxl
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# シートネームがない時の辞書
# ユニークな値を出すときNaNもカウントする
# 自動収集できるような関数を作成する

# csv読み込み時にstrで読み込みたい辞書
kw_dtype = {
    "品番": str,
    "図番": str,
    "元品番": str,
    "先品番": str,
}

kw_encoding = None


def _tk_instance():
    """tkinterのインスタンスを作成し、rootを返します。"""

    root = tk.Tk()
    root.withdraw()

    # topmost指定
    root.attributes("-topmost", True)
    root.withdraw()
    root.lift()
    root.focus_force()

    return root


def get_file_path(file_type=None):
    """
    ファイル選択ダイアログを表示し、ユーザーが選択したファイルのパスを取得します。

    Args:
        file_type (str, optional): 選択可能なファイルの種類。例えば、"txt"を指定すると.txtファイルのみが選択可能になります。
        file_typeがNoneまたは指定されていない場合、全ての種類のファイルが選択可能になります。

    Returns:
        str or None: ユーザーが選択したファイルのパス。ファイルが選択されなかった場合はNoneを返します。
    """

    _tk_instance()

    # ファイル選択ダイアログを表示し、選択したファイルのパスを取得
    if file_type:
        filetypes = ((file_type, f"*.{file_type}"),)
    else:
        filetypes = (("All files", "*.*"),)

    file_path = filedialog.askopenfilename(
        initialdir="..", title="Select file", filetypes=filetypes
    )

    # ファイルが選択されていない場合はNoneを返す
    if not file_path:
        print("ファイルが選択されていません")
        return None

    return file_path


def get_folder_path():
    """
    フォルダ選択ダイアログを表示し、ユーザーが選択したフォルダのパスを取得します。

    Returns:
        str or None: ユーザーが選択したフォルダのパス。フォルダが選択されなかった場合はNoneを返します。
    """

    _tk_instance()

    # フォルダ選択ダイアログを表示し、選択したフォルダのパスを取得
    folder_path = filedialog.askdirectory(initialdir="..", title="Select folder")

    # フォルダが選択されていない場合はNoneを返す
    if not folder_path:
        print("フォルダが選択されていません")
        return None

    return folder_path


# def save_file_path(defaultextension):
#     _tk_instance()

#     # ファイル選択ダイアログを表示し、ファイルを保存するパスを取得
#     file_path = filedialog.asksaveasfilename(
#         initialdir="./", title="Save as", defaultextension=defaultextension
#     )

#     return file_path


def getdf_csv(file_path=None):
    """
    指定されたcsvファイルをpandas DataFrameに読み込みます。file_pathが指定されていない場合は、ファイル選択ダイアログが表示されます。

    Args:
        file_path (str, optional): csvファイルのパス。指定されていない場合、ファイル選択ダイアログが表示されます。

    Returns:
        DataFrame: csvファイルの内容を表すDataFrame。
    """

    # file_pathが指定されていない場合はファイル選択ダイアログを開く
    if file_path is None:
        file_path = get_file_path("csv")

    df = pd.read_csv(file_path, dtype=kw_dtype, encoding=kw_encoding)

    return df


def getdf_xlsx(file_path=None):
    """
    指定されたxlsxファイルをpandas DataFrameに読み込みます。file_pathが指定されていない場合は、ファイル選択ダイアログが表示されます。

    Args:
        file_path (str, optional): xlsxファイルのパス。指定されていない場合、ファイル選択ダイアログが表示されます。

    Returns:
        dict: キーがシート名、値が各シートの内容を表すDataFrameの辞書。
    """
    # file_pathが指定されていない場合はファイル選択ダイアログを開く
    if file_path is None:
        file_path = get_file_path("xlsx")

    # 選択されたファイルをpandasで読み込む。品番をゼロ埋めするためstrで読み込む。
    xls = pd.ExcelFile(file_path)
    dfs = {
        sheet_name: xls.parse(sheet_name, dtype=kw_dtype)
        for sheet_name in xls.sheet_names
    }

    return dfs


def getdf_xlsx_test(file_path=None):
    """
    指定されたxlsxファイルをpandas DataFrameに読み込みます。file_pathが指定されていない場合は、ファイル選択ダイアログが表示されます。

    Args:
        file_path (str, optional): xlsxファイルのパス。指定されていない場合、ファイル選択ダイアログが表示されます。

    Returns:
        dict: キーがシート名、値が各シートの内容を表すDataFrameの辞書。
    """
    # file_pathが指定されていない場合はファイル選択ダイアログを開く
    if file_path is None:
        file_path = get_file_path("xlsx")

    # 選択されたファイルをpandasで読み込む。品番をゼロ埋めするためstrで読み込む。
    xls = pd.ExcelFile(file_path)
    dfs = {
        sheet_name: xls.parse(sheet_name, dtype=kw_dtype)
        for sheet_name in xls.sheet_names
    }

    # 辞書のキー（シート名）の一覧をプリント
    print(f"シート名の一覧: {list(dfs.keys())}")

    return dfs


def get_dataframe(file_path=None, file_type="csv"):
    """
    指定されたファイルをpandas DataFrameに読み込みます。file_pathが指定されていない場合は、ファイル選択ダイアログが表示されます。

    Args:
        file_path (str, optional): ファイルのパス。指定されていない場合、ファイル選択ダイアログが表示されます。
        file_type (str, optional): ファイルの種類。"csv"または"xlsx"。デフォルトは"csv"。

    Returns:
        DataFrameまたは辞書: csvファイルの場合はDataFrame、xlsxファイルの場合はシート名とDataFrameの辞書。
    """
    if file_path is None:
        file_path = get_file_path(file_type)

    if file_type == "csv":
        df = pd.read_csv(file_path, dtype=kw_dtype, encoding=kw_encoding)
        return df

    elif file_type == "xlsx":
        xls = pd.ExcelFile(file_path)
        dfs = {
            sheet_name: xls.parse(sheet_name, dtype=kw_dtype)
            for sheet_name in xls.sheet_names
        }
        return dfs

    else:
        raise ValueError("Unsupported file_type. Choose from 'csv' or 'xlsx'.")


def writedf_xlsx(df_dict):
    """
    引数で指定された辞書に格納されたデータフレームを、指定されたExcelファイルに書き込みます。

    Args:
        df_dict (dict): シート名をキーとし、対応するデータフレームを値とする辞書。
        file_path (str): データを保存するExcelファイルのパス。

    Note:
        辞書の各エントリーに対応するExcelのシートが作成され、そのシートに対応するデータフレームのデータが書き込まれます。
        すでに同名のシートが存在する場合は、そのシートは上書きされます。
        また、データフレームのインデックスはExcelに書き込まれません。
    """
    _tk_instance()

    file_name = filedialog.asksaveasfilename(
        initialdir="..", title="Save as", defaultextension="xlsx"
    )

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def writedf_xlsx2(df_dict, column_angle=0, fit_columns=True, filter_on=True):
    """
    引数で指定された辞書に格納されたデータフレームを、指定されたExcelファイルに書き込みます。

    Args:
        df_dict (dict): シート名をキーとし、対応するデータフレームを値とする辞書。
        column_angle (int, optional): カラムの文字列の向きを表す角度。デフォルトは0。
        fit_columns (bool, optional): 列の幅を値に合わせて調整するかどうか。デフォルトはTrue。
        filter_on (bool, optional): カラムのフィルタを有効にするかどうか。デフォルトはTrue。

    Note:
        辞書の各エントリーに対応するExcelのシートが作成され、そのシートに対応するデータフレームのデータが書き込まれます。
        すでに同名のシートが存在する場合は、そのシートは上書きされます。
        また、データフレームのインデックスはExcelに書き込まれません。
    """
    _tk_instance()

    file_name = filedialog.asksaveasfilename(
        initialdir="..", title="Save as", defaultextension="xlsx"
    )

    for sheet_name, df in df_dict.items():
        df.columns = df.columns.astype(str)  # 列見出しを文字列に変換
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        if filter_on:
            tab = Table(
                displayName=sheet_name, ref=f"A1:{chr(65 + len(df.columns))}{len(df)+1}"
            )
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

        for cell in ws[1]:
            cell.alignment = Alignment(textRotation=column_angle)

        if fit_columns:
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(file_name)


def writedf_xlsx3(df_dict, angle=0, adjust_col_width=False, enable_filter=False):
    """
    引数で指定された辞書に格納されたデータフレームを、指定されたExcelファイルに書き込みます。

    Args:
        df_dict (dict): シート名をキーとし、対応するデータフレームを値とする辞書。
        angle (int): カラムの文字列の方向の角度。デフォルトは0（水平）。
        adjust_col_width (bool): 列の幅を値にフィットするかどうか。デフォルトはFalse。
        enable_filter (bool): カラム（1行目）のフィルタをONにするかどうか。デフォルトはFalse。

    Note:
        辞書の各エントリーに対応するExcelのシートが作成され、そのシートに対応するデータフレームのデータが書き込まれます。
        すでに同名のシートが存在する場合は、そのシートは上書きされます。
        また、データフレームのインデックスはExcelに書き込まれません。
    """
    _tk_instance()

    file_name = filedialog.asksaveasfilename(
        initialdir="..", title="Save as", defaultextension="xlsx"
    )

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            if angle != 0 or adjust_col_width or enable_filter:
                # Load workbook
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                if enable_filter:
                    worksheet.auto_filter.ref = worksheet.dimensions

                for idx, column in enumerate(df.columns, start=1):
                    cell = worksheet.cell(row=1, column=idx)
                    if angle != 0:
                        cell.alignment = openpyxl.styles.Alignment(textRotation=angle)

                    if adjust_col_width:
                        column_width = max(
                            (len(str(cell_value)) for cell_value in df[column])
                        )
                        worksheet.column_dimensions[
                            openpyxl.utils.get_column_letter(idx)
                        ].width = column_width


def csv_to_xlsx():
    """_summary_"""

    file_path = get_file_path("csv")
    df = pd.read_csv(file_path, dtype=kw_dtype, encoding=kw_encoding)
    df_dict = {"sheet1": df}
    writedf_xlsx(df_dict)


def csv_to_excel(output_file_name="output.xlsx"):
    """
    指定したディレクトリ内の全てのCSVファイルを読み込み、それらを一つのExcelファイルに変換します。

    この関数は、デフォルトで"output.xlsx"という名前のExcelファイルを作成しますが、
    出力するファイル名はパラメータで変更することができます。各CSVファイルはExcelファイルの
    異なるシートに書き込まれ、シート名は元のCSVファイル名となります。

    Parameters
    ----------
    output_file_name : str, optional
        出力するExcelファイルの名前、デフォルトは "output.xlsx"

    Returns
    -------
    None
    """

    directory = get_folder_path()

    with pd.ExcelWriter(output_file_name, engine="openpyxl") as writer:
        for filename in os.listdir(directory):
            if filename.endswith(".csv"):
                file_path = os.path.join(directory, filename)
                df = pd.read_csv(
                    os.path.join(directory, filename),
                    dtype=kw_dtype,
                    encoding=kw_encoding,
                )
                df.excel(writer, sheet_name=os.path.splitext(filename)[0], index=False)
